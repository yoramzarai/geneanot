# pylint: disable=line-too-long,invalid-name,pointless-string-statement,too-many-arguments,too-many-lines,too-many-instance-attributes,too-many-locals
# type: ignore   # for Pylance
"""
Utils related to excel spreadsheet.

This requires also xlsxwriter to be installed.
"""
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

def dfs_to_excel_file(dfs: list[pd.DataFrame], excel_file_name: str, sheet_names: list[str],
                      add_index: bool = False,
                      na_rep: str = 'NaN',
                      float_format: str | None = None,
                      extra_width: int = 0,
                      header_format: dict | None = None):
    """
    Write DataFrames to excel, while auto adjusting column widths based on the data.
    """
    if len(dfs) != len(sheet_names):
        raise ValueError("dfs_to_excel_file: the numbers of dfs and sheet names must match !!")

    with pd.ExcelWriter(excel_file_name) as writer:
        for sheet_name, df in zip(sheet_names, dfs):
            df.to_excel(writer, sheet_name=sheet_name, index=add_index, na_rep=na_rep, float_format=float_format)

            # Auto-adjust columns' width
            worksheet = writer.sheets[sheet_name]
            for column in df:
                #column_width = max(df[column].astype(str).map(len).max(), len(column)) + extra_width  # this (.map(len)) does not work for NA values
                #column_width = max(df[column].fillna(method='ffill').astype(str).map(len).max(), len(column)) + extra_width  # this first replaces NA values with previous valid value from the column
                column_width = max(df[df[column].notna()][column].astype(str).map(len).max(), len(column)) + extra_width  # this first removes NA values from the column
                col_idx = df.columns.get_loc(column)
                worksheet.set_column(col_idx, col_idx, column_width)

            if header_format is not None:
                h_format = writer.book.add_format(header_format)
                for col_num, value in enumerate(df.columns.values):
                    worksheet.write(0, col_num, value, h_format)

def create_excel_description_sheet(excel_file: str, info: dict, sheet_name: str = "Desc", sheet_pos: int | None = None) -> bool:
    """
    This function creates a sheet containing the key [value] of info in column 1 [2] of the sheet.
    If a key is empty (i.e., k==''), an empty line is created.

    sheet_pos: the position of the sheet in the file. 0 for the first sheet, None for the last sheet.
    """
    wb = load_workbook(excel_file)
    wb.create_sheet(sheet_name, sheet_pos)
    sheet = wb[sheet_name]
    for i, (k, v) in enumerate(info.items(), start=1):
        if k != "":
            sheet.cell(row=i, column=1).value = k
            sheet.cell(row=i, column=1).font = Font(bold=True)

            sheet.cell(row=i, column=2).value = v

    # set column width
    sheet.column_dimensions["A"].width = max([len(x) for x in info.keys()])
    sheet.column_dimensions["B"].width = max([len(x) for x in info.values()])

    wb.save(excel_file)
    return True
